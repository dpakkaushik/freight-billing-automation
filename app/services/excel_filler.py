"""Fill the Swaraj Gujarat invoice template with one or more LR rows.

Cell map (agreed with user):

    Header area
        I14    "Invoice No : PTLM-2627SWM-{suffix}"
        I15    "Invoice Date : dd/mm/yyyy"

    Data rows (one per Delivery Order No., starting at row 27, max 12 rows)
        C{row}    delivery_order_no
        D{row}    gcn_no
        E{row}    gcn_date
        F{row}    vehicle_no
        G{row}    destination       ("To" city from LR)
        H{row}    delivery_date     (hand-written on LR, typed by user)
        I{row}    qty
        J{row}    "Fixed"           (literal text)
        L{row}    total_amount

    Totals
        L39      formula updated to =SUM(L27:L38) if more than one row used
        B41      replaced with Indian amount-in-words (template's external
                 SpellCurr macro can't be resolved by LibreOffice)

The template is loaded from app/templates/swaraj_invoice_gujarat.xlsx.
"""
from __future__ import annotations

import re
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any

from loguru import logger
from openpyxl import load_workbook

from app.services.amount_words import amount_in_words


# Located inside the package so it ships with the code
TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "templates" / "swaraj_invoice_gujarat.xlsx"

SHEET_NAME = "Gujarat sawraj"

# Data rows available in the template (S.No. 1..12)
DATA_FIRST_ROW = 27
DATA_LAST_ROW = 38


def _format_invoice_no(suffix: str) -> str:
    return f"Invoice No : PTLM-2627SWM-{suffix}"


def _format_invoice_date(d: date) -> str:
    return f"Invoice Date : {d.strftime('%d/%m/%Y')}"


def _to_decimal(amount: Any) -> Decimal:
    if amount in (None, ""):
        return Decimal("0")
    if isinstance(amount, Decimal):
        return amount
    return Decimal(str(amount).replace(",", "").strip() or "0")


def _sum_qty(lrs: list[dict[str, Any]]) -> int:
    total = 0
    for lr in lrs:
        m = re.search(r"(\d+)", str(lr.get("qty") or ""))
        if m:
            total += int(m.group(1))
    return total


def build_rows(lrs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Flatten confirmed LR list into Excel rows using the stacking rule.

    Single LR (image upload):
      One row per DO. Each row gets all LR fields + its own total_amount.

    Multiple LRs (PDF upload — up to 3 GCNs):
      Stacking rule agreed with user:
        - Column C (DO#):         all DOs from all LRs stacked one under another
        - Column D (LR/GCN No):   all GCN numbers stacked (one per LR)
        - Column E (LR Date):     first row only (common)
        - Column F (Vehicle No):  first row only (common)
        - Column G (Destination): all destinations stacked (one per LR)
        - Column H (Del. Date):   first row only (common)
        - Column I (Qty):         first row only — combined total across all LRs
        - Column L (Amount):      first row only — freight amount entered by user
    """
    if not lrs:
        return []

    if len(lrs) == 1:
        # Single LR — original per-DO behaviour preserved
        lr = lrs[0]
        do_rows = lr.get("do_rows") or [{"delivery_order_no": "", "total_amount": ""}]
        rows = []
        for idx, do in enumerate(do_rows):
            first = idx == 0
            rows.append({
                "delivery_order_no": (do.get("delivery_order_no") or "").strip(),
                "gcn_no":        (lr.get("gcn_no") or "").strip(),
                "gcn_date":      (lr.get("gcn_date") or "").strip() if first else "",
                "vehicle_no":    (lr.get("vehicle_no") or "").strip() if first else "",
                "destination":   (lr.get("destination") or "").strip() if first else "",
                "delivery_date": (lr.get("delivery_date") or "").strip() if first else "",
                "qty":           (lr.get("qty") or "").strip() if first else "",
                "total_amount":  (do.get("total_amount") or "").strip(),
            })
        return rows

    # Multiple LRs (PDF) — stacking rule
    all_dos: list[dict] = []
    for lr in lrs:
        for do in (lr.get("do_rows") or [{"delivery_order_no": "", "total_amount": ""}]):
            all_dos.append(do)

    gcn_numbers  = [(lr.get("gcn_no") or "").strip() for lr in lrs]
    destinations = [(lr.get("destination") or "").strip() for lr in lrs]

    first_lr = lrs[0]
    common_gcn_date      = (first_lr.get("gcn_date") or "").strip()
    common_vehicle       = (first_lr.get("vehicle_no") or "").strip()
    common_delivery_date = (first_lr.get("delivery_date") or "").strip()
    total_qty = _sum_qty(lrs)
    qty_display = f"{total_qty} Units" if total_qty > 0 else (first_lr.get("qty") or "").strip()

    # Freight amount: take from the first DO of the first LR
    first_amount = (all_dos[0].get("total_amount") or "").strip() if all_dos else ""

    rows = []
    for i, do in enumerate(all_dos):
        rows.append({
            "delivery_order_no": (do.get("delivery_order_no") or "").strip(),
            "gcn_no":        gcn_numbers[i]  if i < len(gcn_numbers)  else "",
            "gcn_date":      common_gcn_date      if i == 0 else "",
            "vehicle_no":    common_vehicle        if i == 0 else "",
            "destination":   destinations[i] if i < len(destinations) else "",
            "delivery_date": common_delivery_date  if i == 0 else "",
            "qty":           qty_display           if i == 0 else "",
            "total_amount":  first_amount          if i == 0 else "",
        })
    return rows


def fill_invoice(
    suffix: str,
    invoice_date: date,
    lrs: list[dict[str, Any]],
    output_path: Path,
) -> Path:
    """Open the template, write all fields, save to `output_path`.

    Returns the output path.
    Raises ValueError if the template is missing or the LR list overflows
    the available row count.
    """
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Excel template not found at {TEMPLATE_PATH}")

    rows = build_rows(lrs)
    max_rows = DATA_LAST_ROW - DATA_FIRST_ROW + 1
    if len(rows) > max_rows:
        raise ValueError(
            f"Template supports at most {max_rows} rows but {len(rows)} were given. "
            "Split into multiple invoices."
        )

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    # --- Header ---
    ws["I14"] = _format_invoice_no(suffix)
    ws["I15"] = _format_invoice_date(invoice_date)

    # --- Data rows ---
    total_amount = Decimal("0")
    rows_used = max(1, len(rows))
    for i, row_data in enumerate(rows):
        r = DATA_FIRST_ROW + i
        ws[f"C{r}"] = row_data["delivery_order_no"]
        ws[f"D{r}"] = row_data["gcn_no"]
        ws[f"E{r}"] = row_data["gcn_date"]
        ws[f"F{r}"] = row_data["vehicle_no"]
        ws[f"G{r}"] = row_data["destination"]
        ws[f"H{r}"] = row_data["delivery_date"]
        ws[f"I{r}"] = row_data["qty"]
        ws[f"J{r}"] = "Fixed"
        amt = _to_decimal(row_data["total_amount"])
        # Write as a number if it parsed cleanly; else leave the raw string.
        if amt > 0:
            ws[f"L{r}"] = float(amt)
            total_amount += amt
        elif row_data["total_amount"]:
            ws[f"L{r}"] = row_data["total_amount"]

    # --- Total formula (L39) ---
    # Template ships with =+L27 which only covers one row. Replace with a
    # SUM over the rows we actually used so multi-row invoices total correctly.
    if rows_used > 1:
        ws["L39"] = f"=SUM(L{DATA_FIRST_ROW}:L{DATA_FIRST_ROW + rows_used - 1})"
    # else: leave the existing =+L27 formula in place.

    # --- Amount in words (B41) ---
    # The template formula =[1]!SpellCurr(L41) calls a macro in another
    # workbook that LibreOffice can't resolve. Compute it here.
    igst_18 = total_amount * Decimal("0.18")
    grand_total = total_amount + igst_18
    ws["B41"] = amount_in_words(grand_total)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Wrote invoice xlsx: {} ({} rows)", output_path, rows_used)
    return output_path
